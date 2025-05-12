import torch
import numpy as np
import pandas as pd
from data_process_text import load_all_code_ids
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
workload_dict = {
    "Spark ConnectedComponent Application": "CC",
    "DecisionTree classification Example": "DT",
    "Spark KMeans Example": "KM",
    "LinerRegressionApp Example": "LiR",
    "LogisticRegressionApp Example": "LoR",
    "Spark LabelPropagation Application": "LP",
    "MFApp Example": "MF",
    "Spark PCA Example": "PCA",
    "Spark PregelOperation Application": "PO",
    "Spark PageRank Application": "PR",
    "Spark StronglyConnectedComponent Application": "SCC",
    "Spark ShortestPath Application": "SP",
    "Spark SVDPlusPlus Application": "SVD",
    "SVM Classifier Example": "SVM",
    "Spark TriangleCount Application": "TC",
    "TeraSort": "TS"
}

#2024-12-3  以stage的运行时间为预测对象
def read_stage_dataset(dataset_path):
    i=0
    j=0
    ws2nodes = torch.load('dag_data/ws2nodes.pth')
    ws2adj = torch.load('dag_data/ws2adj.pth')
    all_code_dict = load_all_code_ids()
    df = pd.read_csv(dataset_path, sep=',', low_memory=False)
    workload, stage_id = df['AppName'].apply(lambda x: workload_dict[x]), df['stage_id']
    code_vec = []
    nodes_vec = []
    adj_vec = []
    W, S = [], []
    for w, s in zip(workload, stage_id):
        W.append(w)
        S.append(s)
        #本try,except结构是为了找出当前stage对应的向量
        try:
            code_vec.append(all_code_dict[w][str(s)])
            i=i+1
        except:
            code_vec.append([0 for _ in range(1000)])
            #code_vec.append([1 for _ in range(1000)])
        #本try,except
        try:
            nodes_vec.append(ws2nodes[w][str(s)])
            adj_vec.append(ws2adj[w][str(s)])
            j=j+1
        except:
            nodes_vec.append(np.zeros(64, dtype=float))
            adj_vec.append(np.zeros([64, 64], dtype=float))
            #nodes_vec.append(np.ones(64, dtype=float))
            #adj_vec.append(np.ones([64, 64], dtype=float))
    Y = df['duration']
    #Y= df['Duration']/1000
    X_df = df.drop(['AppId', 'AppName', 'Duration', 'code', 'duration'], axis=1)
    X = X_df.values
    print('code不为0的数目:')
    print(i)
    print('adj不为0的数目:')
    print(j)
    #X为配置参数，stage的输入输出读写，当前环境的特征，应用特征；
    #code_vec为当前stage在运行中调用的代码
    #Y为stage的运行时间
    #W为workload的名字
    #S为stage的ID
    #nodes_vec为
    #adj_vec为与当前stage相连的stage
    return X, code_vec, Y, W, S, nodes_vec, adj_vec

#V1-2025-1-5 数据集以workload的运行时间为预测对象;
def read_workload_dataset(dataset_path):
    i=0
    j=0
    ws2nodes = torch.load('dag_data/ws2nodes.pth',weights_only=False)
    ws2adj = torch.load('dag_data/ws2adj.pth',weights_only=False)
    all_code_dict = load_all_code_ids()
    df = pd.read_csv(dataset_path, sep=',', low_memory=False)
    workload, stage_id = df['AppName'].apply(lambda x: workload_dict[x]), df['stage_id']
    code_vec = []
    nodes_vec = []
    adj_vec = []
    W, S = [], []
    for w, s in zip(workload, stage_id):
        W.append(w)
        S.append(s)
        #本try,except结构是为了找出当前stage对应的向量
        try:
            code_vec.append(all_code_dict[w][str(s)])
            i=i+1
        except:
            code_vec.append([0 for _ in range(1000)])
            #code_vec.append([1 for _ in range(1000)])
        #本try,except
        try:
            nodes_vec.append(ws2nodes[w][str(s)])
            adj_vec.append(ws2adj[w][str(s)])
            j=j+1
        except:
            nodes_vec.append(np.zeros(64, dtype=float))
            adj_vec.append(np.zeros([64, 64], dtype=float))
            #nodes_vec.append(np.ones(64, dtype=float))
            #adj_vec.append(np.ones([64, 64], dtype=float))
    #Y = df['duration']
    Y= df['Duration']/1000
    X_df = df.drop(['AppId', 'AppName', 'Duration', 'code', 'duration', 'stage_id'], axis=1)
    X = X_df.values
    print('code不为0的数目:')
    print(i)
    print('adj不为0的数目:')
    print(j)
    #X为配置参数，stage的输入输出读写，当前环境的特征，应用特征；
    #code_vec为当前stage在运行中调用的代码
    #Y为stage的运行时间
    #W为workload的名字
    #S为stage的ID
    #nodes_vec为
    #adj_vec为与当前stage相连的stage
    return X, code_vec, Y, W, S, nodes_vec, adj_vec


def normalize_column(column, min_val, max_val):
    if min_val == max_val:
        return 0
    else:
        return (column - min_val) / (max_val - min_val)

def update_conf_range(conf_range_path, updated_conf_range):
    book = load_workbook(conf_range_path)
    sheet = book.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None
    for r_idx, row in enumerate(dataframe_to_rows(updated_conf_range, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    book.save(conf_range_path)

#2024-12-3 构建混合粒度的数据集
def build_mixed_dataset(dataset_path, scaler_save_path, conf_range_path):
    i = 0
    j = 0
    ws2nodes = torch.load('dag_data/updated/ws2nodes.pth', weights_only=False)
    ws2adj = torch.load('dag_data/updated/ws2adj.pth', weights_only=False)
    all_code_dict = load_all_code_ids()
    df = pd.read_csv(dataset_path, sep=',', low_memory=False)
    
    # 读取配置参数范围

    conf_range_df = pd.read_excel(conf_range_path)

    conf_range_dict = conf_range_df.set_index('Parameter').T.to_dict('list')
    
    # 选择数值列
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    numeric_cols = numeric_cols.drop(['Duration', 'stage_id'])  # 排除 'Duration' 列
    df.fillna(0, inplace=True)
    # 对数据进行归一化，并更新配置参数范围
    updated_conf_range = conf_range_df.copy()
    for col in numeric_cols:
        if col in conf_range_dict:
            min_val, max_val = conf_range_dict[col][1], conf_range_dict[col][2]
            col_min, col_max = df[col].min(), df[col].max()
            if col_min < min_val:
                min_val = col_min
            if col_max > max_val:
                max_val = col_max
            updated_conf_range.loc[updated_conf_range['Parameter'] == col, ['min', 'max']] = [min_val, max_val]
        else:
            min_val, max_val = df[col].min(), df[col].max()
            new_row = pd.DataFrame({'Parameter': [col], 'min': [min_val], 'max': [max_val]})
            updated_conf_range = pd.concat([updated_conf_range, new_row], ignore_index=True)
        df[col] = normalize_column(df[col].fillna(0), min_val, max_val)
    
    # 保存更新后的配置参数范围
    update_conf_range(conf_range_path, updated_conf_range)
    
    # 保存最大值和最小值
    normalization_params = {'min_vals': df[numeric_cols].min(), 'max_vals': df[numeric_cols].max()}
    torch.save(normalization_params, scaler_save_path)
    
    # 将 Duration 的单位转换为秒
    df['Duration'] = df['Duration'].apply(lambda x: x / 1000)

    code_vec = {m: [] for m in range(len(df['AppId'].unique()))}
    nodes_vec = {m: [] for m in range(len(df['AppId'].unique()))}
    adj_vec = {m: [] for m in range(len(df['AppId'].unique()))}
    stage_features = {m: [] for m in range(len(df['AppId'].unique()))}
    workload_features = {m: [] for m in range(len(df['AppId'].unique()))}
    Y = {m: [] for m in range(len(df['AppId'].unique()))}
    ys = {m: [] for m in range(len(df['AppId'].unique()))}
    W, S = [], []
    m = 0
    # 通过 APPID 分组，整合出内部数据
    groups = df.groupby('AppId')

    for app_id, group in groups:
        workload, stage_id = group['AppName'].apply(lambda x: workload_dict[x]), group['stage_id']
        for w, s in zip(workload, stage_id):
            S.append(s)
            # 本 try, except 结构是为了找出当前 stage 对应的向量
            try:
                code_vec[m].append(all_code_dict[w][str(s)])
                i = i + 1
            except:
                code_vec[m].append([0 for _ in range(1000)])
            try:
                nodes_vec[m].append(ws2nodes[w][str(s)])
                adj_vec[m].append(ws2adj[w][str(s)])
                j = j + 1
            except:
                nodes_vec[m].append(np.zeros(64, dtype=float))
                adj_vec[m].append(np.zeros([64, 64], dtype=float))
        
        W.append(w)
        #去掉stage_id
        #stage_feature = group[['stage_id', 'input', 'output', 'read', 'write']]
        stage_feature = group[['input', 'output', 'read', 'write']]
        y = group['Duration']
        y_first_row = y.iloc[0]
        Y[m] = y_first_row
        ys[m] = group['duration'].values
        stage_feature = stage_feature.values
        workload_feature = group.drop(['stage_id', 'duration', 'input', 'output', 'read', 'write', 'AppId', 'AppName', 'Duration', 'code'], axis=1)
        # 只保留第一行数据
        workload_features_first_row = workload_feature.iloc[0]
        workload_features_first_row = workload_features_first_row.values
        #保留全部数据看看效果
        #workload_features_first_row = workload_feature.values

        stage_features[m] = stage_feature
        workload_features[m] = workload_features_first_row
        m = m + 1

    print('code不为0的数目:')
    print(i)
    print('adj不为0的数目:')
    print(j)
    return W, S, code_vec, nodes_vec, adj_vec, stage_features, workload_features, Y, ys


if __name__ == '__main__':
    machine_num = '3'
    #X, CODE, Y, W, S, NODES, ADJS = read_stage_dataset(dataset_path='test_data/dataset_test_'+machine_num+'.csv')
    #all_data
    #X, CODE, Y, W, S, NODES, ADJS = read_stage_dataset(dataset_path='dataset/dataset_4/train/train_data.csv')
    #X, CODE, Y, W, S, NODES, ADJS = read_stage_dataset(dataset_path='dataset/dataset_4/test/test_data.csv')
    #never-seen
    X, CODE, Y, W, S, NODES, ADJS = read_stage_dataset(dataset_path='dataset/dataset_1/never_seen/cold-start/train/train_data_S.csv')
    #X, CODE, Y, W, S, NODES, ADJS = read_stage_dataset(dataset_path='dataset/dataset_1/never_seen/cold-start/test/test_data_NS.csv')
    #trans-learn
    #X, CODE, Y, W, S, NODES, ADJS = read_stage_dataset(dataset_path='dataset/dataset_3/trans_learn/test/test_data_NS.csv')
    #X, CODE, Y, W, S, NODES, ADJS = read_stage_dataset(dataset_path='dataset/dataset_3/trans_learn/train/train_data_S.csv')
    #X, CODE, Y, W, S, NODES, ADJS = read_workload_dataset(dataset_path='dataset/dataset_1-4/train/train_data.csv')


    #W,S,CODE, NODES,ADJS,Stage_Features,Workload_Features,Y = build_mixed_dataset(dataset_path='dataset/dataset_2/train/train_data.csv')
    print(X.shape)
    print(len(CODE))
    print(len(Y))
    print(len(NODES))
    print(len(ADJS))
    print(len(W))
    print(len(S))
    #W为负载类型，S为stage编号，X为配置参数 ，Y为stage的运行时间
    np.save('dataset/X.npy', X)
    np.save('dataset/CODE.npy', CODE)
    np.save('dataset/Y.npy', Y)
    np.save('dataset/NODES.npy', NODES)
    np.save('dataset/ADJS.npy', ADJS)
    np.save('dataset/W.npy', W)
    np.save('dataset/S.npy', S)
    #np.save('dataset/Stage_Features.npy', Stage_Features)
    #np.save('dataset/Workload_Features.npy', Workload_Features)

    #np.save('dataset/dataset_4/X.npy', X)
    #np.save('dataset/dataset_4/CODE.npy', CODE)
    #np.save('dataset/dataset_4/Y.npy', Y)
    #np.save('dataset/dataset_4/NODES.npy', NODES)
    #np.save('dataset/dataset_4/ADJS.npy', ADJS)
    #np.save('dataset/dataset_4/W.npy', W)
    #np.save('dataset/dataset_4/S.npy', S)
    #never-seen
    #np.save('dataset/dataset_3/trans_learn/test/X.npy', X)
    #np.save('dataset/dataset_3/trans_learn/test/CODE.npy', CODE)
    #np.save('dataset/dataset_3/trans_learn/test/Y.npy', Y)
    #np.save('dataset/dataset_3/trans_learn/test/NODES.npy', NODES)
    #np.save('dataset/dataset_3/trans_learn/test/ADJS.npy', ADJS)
    #np.save('dataset/dataset_3/trans_learn/test/W.npy', W)
    #np.save('dataset/dataset_3/trans_learn/test/S.npy', S)
